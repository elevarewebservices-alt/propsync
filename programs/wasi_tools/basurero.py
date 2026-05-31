"""
WASI - Inactivar, Comentar y Enviar a Papelera desde Excel
===========================================================
Uso:
    python archivar_excel.py <ARCHIVO.xlsx>

Ejemplo:
    python archivar_excel.py propiedades.xlsx

El Excel debe tener estas columnas:
    - "Disponible (codigo)"  → ID de la propiedad
    - "Comentario"           → Comentario a agregar
"""

import requests
import sys
import time

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ─── CREDENCIALES ────────────────────────────────────────────────────────────
ID_COMPANY  = "2946576"
WASI_TOKEN  = "WGOj_hXgB_xRY4_dQOy"
BASE_URL    = "https://api.wasi.co/v1"

# ─── FUNCIONES ───────────────────────────────────────────────────────────────
def credenciales():
    return {"id_company": ID_COMPANY, "wasi_token": WASI_TOKEN}

def get_propiedad(id_property):
    url = f"{BASE_URL}/property/get/{id_property}"
    resp = requests.post(url, data=credenciales())
    resp.raise_for_status()
    data = resp.json()
    if data.get("status") == "error":
        raise Exception(f"Error Wasi: {data.get('message')}")
    return data

def post_update(id_property, campos):
    url = f"{BASE_URL}/property/update/{id_property}"
    data = {**credenciales(), **campos}
    resp = requests.post(url, data=data)
    resp.raise_for_status()
    resultado = resp.json()
    if resultado.get("status") == "error":
        raise Exception(f"Error: {resultado.get('message')}")
    return resultado

def leer_excel(archivo):
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        col_id  = headers.index("Disponible (codigo)") + 1
        col_com = headers.index("Comentario") + 1
    except ValueError as e:
        raise Exception(f"No se encontró la columna: {e}. Columnas: {headers}")

    propiedades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_prop    = row[col_id - 1]
        comentario = row[col_com - 1]
        if id_prop:
            propiedades.append((str(int(id_prop)), str(comentario) if comentario else ""))
    return propiedades

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Uso: python archivar_excel.py <ARCHIVO.xlsx>")
        sys.exit(1)

    archivo = sys.argv[1]
    print(f"\n{'='*55}")
    print(f"  WASI - Inactivar y Archivar desde Excel")
    print(f"{'='*55}")

    try:
        print(f"\nLeyendo archivo: {archivo}...")
        propiedades = leer_excel(archivo)
        print(f"✓ {len(propiedades)} propiedades encontradas\n")
    except FileNotFoundError:
        print(f"\n✗ No se encontró el archivo: {archivo}")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Error leyendo Excel: {e}")
        sys.exit(1)

    exitosos = 0
    fallidos = 0
    errores  = []

    for i, (id_property, comentario) in enumerate(propiedades, 1):
        print(f"\n[{i}/{len(propiedades)}] Propiedad #{id_property}...")
        try:
            propiedad = get_propiedad(id_property)
            titulo = propiedad.get("title", id_property)
            print(f"      ✓ Encontrada: {titulo}")

            # Paso 1 - Inactivar y comentar
            print(f"      → Inactivando y comentando...")
            post_update(id_property, {
                "id_status_on_page": "2",
                "comment": comentario
            })
            print(f"      ✓ Inactiva + Comentario agregado")
            time.sleep(1)

            # Paso 2 - Enviar a papelera (id_status_on_page: 4 = Deleted)
            print(f"      → Enviando a papelera...")
            post_update(id_property, {"id_status_on_page": "4"})
            print(f"      ✓ Enviada a papelera")

            exitosos += 1

        except Exception as e:
            print(f"      ✗ Error: {e}")
            errores.append((id_property, str(e)))
            fallidos += 1

        if i < len(propiedades):
            time.sleep(1)

    print(f"\n{'='*55}")
    print(f"  ¡Proceso completado!")
    print(f"  ✓ Exitosos: {exitosos}")
    print(f"  ✗ Fallidos: {fallidos}")
    if errores:
        print(f"\n  Propiedades con error:")
        for id_prop, err in errores:
            print(f"    - #{id_prop}: {err}")
    print(f"{'='*55}\n")

if __name__ == "__main__":
    main()