"""
WASI - Inactivar Propiedades desde Excel
=========================================
Uso:
    python inactivar_excel.py <ARCHIVO.xlsx>

Ejemplo:
    python inactivar_excel.py propiedades.xlsx

El Excel debe tener estas columnas:
    - "Disponible (codigo)"  → ID de la propiedad
    - "Comentario"           → Comentario a agregar

"""

import requests
import sys
import time

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ─── CREDENCIALES ────────────────────────────────────────────────────────────
ID_COMPANY  = "2946576"
WASI_TOKEN  = "WGOj_hXgB_xRY4_dQOy"
BASE_URL    = "https://api.wasi.co/v1"

# ─── FUNCIONES ───────────────────────────────────────────────────────────────
def credenciales():
    return {"id_company": ID_COMPANY, "wasi_token": WASI_TOKEN}

def get_propiedad(id_property):
    url = f"{BASE_URL}/property/get/{id_property}"
    resp = requests.post(url, data=credenciales())
    resp.raise_for_status()
    data = resp.json()
    if data.get("status") == "error":
        raise Exception(f"Error Wasi: {data.get('message', 'Sin mensaje')}")
    return data

def inactivar_y_comentar(id_property, comentario):
    url = f"{BASE_URL}/property/update/{id_property}"
    payload = credenciales()
    payload["id_status_on_page"] = "2"
    payload["comment"] = comentario
    resp = requests.post(url, data=payload)
    resp.raise_for_status()
    resultado = resp.json()
    if resultado.get("status") == "error":
        raise Exception(f"Error al actualizar: {resultado.get('message')}")
    return resultado

def leer_excel(archivo):
    """Lee el Excel y devuelve lista de (id, comentario)."""
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active

    # Encontrar columnas por nombre
    headers = [cell.value for cell in ws[1]]
    try:
        col_id  = headers.index("Disponible (codigo)") + 1
        col_com = headers.index("Comentario") + 1
    except ValueError as e:
        raise Exception(f"No se encontró la columna: {e}. Columnas disponibles: {headers}")

    propiedades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_prop    = row[col_id - 1]
        comentario = row[col_com - 1]
        if id_prop:  # Ignorar filas vacías
            propiedades.append((str(int(id_prop)), str(comentario) if comentario else ""))

    return propiedades

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Uso: python inactivar_excel.py <ARCHIVO.xlsx>")
        print("Ejemplo: python inactivar_excel.py propiedades.xlsx")
        sys.exit(1)

    archivo = sys.argv[1]
    print(f"\n{'='*55}")
    print(f"  WASI - Inactivar Propiedades desde Excel")
    print(f"{'='*55}")

    try:
        print(f"\nLeyendo archivo: {archivo}...")
        propiedades = leer_excel(archivo)
        print(f"✓ {len(propiedades)} propiedades encontradas\n")
    except FileNotFoundError:
        print(f"\n✗ No se encontró el archivo: {archivo}")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Error leyendo Excel: {e}")
        sys.exit(1)

    exitosos  = 0
    fallidos  = 0
    errores   = []

    for i, (id_property, comentario) in enumerate(propiedades, 1):
        print(f"[{i}/{len(propiedades)}] Propiedad #{id_property}...")
        try:
            # Obtener título
            propiedad = get_propiedad(id_property)
            titulo = propiedad.get("title", id_property)

            # Inactivar y comentar
            inactivar_y_comentar(id_property, comentario)
            print(f"      ✓ {titulo}")
            print(f"      ✓ Comentario: \"{comentario}\"")
            exitosos += 1

        except Exception as e:
            print(f"      ✗ Error: {e}")
            errores.append((id_property, str(e)))
            fallidos += 1

        # Pequeña pausa para no saturar la API
        if i < len(propiedades):
            time.sleep(1)

    print(f"\n{'='*55}")
    print(f"  ¡Proceso completado!")
    print(f"  ✓ Exitosos: {exitosos}")
    print(f"  ✗ Fallidos: {fallidos}")
    if errores:
        print(f"\n  Propiedades con error:")
        for id_prop, err in errores:
            print(f"    - #{id_prop}: {err}")
    print(f"{'='*55}\n")

if __name__ == "__main__":
    main()