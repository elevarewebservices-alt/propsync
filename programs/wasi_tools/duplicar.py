"""
WASI - Duplicar, Activar y Archivar Original desde Excel
=========================================================
Uso:
    python auto_excel.py <ARCHIVO.xlsx>

El Excel debe tener estas columnas:
    - "Disponible (codigo)"  → ID de la propiedad (obligatorio)
    - "Comentario"           → Comentario a agregar a la original antes de archivar (obligatorio)
"""

import requests
import sys
import os
import time
import tempfile

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ─── CREDENCIALES ────────────────────────────────────────────────────────────
ID_COMPANY  = "2946576"
WASI_TOKEN  = "WGOj_hXgB_xRY4_dQOy"
BASE_URL    = "https://api.wasi.co/v1"


CAMPOS_EXCLUIR = {
    "id_property", "id_company", "status", "created_at", "updated_at",
    "visits", "url", "galleries", "main_image", "features", "user_data",
    "link", "id_property_wasi", "published", "trash",
    "country_label", "region_label", "city_label", "location_label",
    "zone_label", "iso_currency", "name_currency", "unit_area_label",
    "unit_built_area_label", "unit_private_area_label", "maintenance_fee_label",
    "sale_price_label", "rent_price_label", "rents_type_label",
    "property_condition_label", "status_on_page_label", "publish_on_map_label",
    "availability_label", "label", "label_color"
}

# ─── FUNCIONES ───────────────────────────────────────────────────────────────
def credenciales():
    return {"id_company": ID_COMPANY, "wasi_token": WASI_TOKEN}

def get_propiedad(id_property):
    url = f"{BASE_URL}/property/get/{id_property}"
    resp = requests.post(url, data=credenciales())
    resp.raise_for_status()
    data = resp.json()
    if data.get("status") == "error":
        raise Exception(f"Error Wasi: {data.get('message')}")
    return data

def get_propietario(id_property):
    url = f"{BASE_URL}/property/owner/{id_property}"
    resp = requests.post(url, data=credenciales())
    resp.raise_for_status()
    data = resp.json()
    if data.get("status") == "error" or data.get("total", 0) == 0:
        return None
    return data

def asignar_propietario(id_property_nueva, id_client):
    url = f"{BASE_URL}/client/add-property/{id_property_nueva}"
    data = {**credenciales(), "id_client": id_client, "id_client_type": "5"}
    resp = requests.post(url, data=data)
    resp.raise_for_status()
    resultado = resp.json()
    if resultado.get("status") == "error":
        raise Exception(f"Error asignando propietario: {resultado.get('message')}")
    return resultado

def get_portales(id_property):
    """Obtiene los portales activos de la propiedad."""
    url = f"{BASE_URL}/portal/property/{id_property}"
    resp = requests.get(url, params=credenciales())
    resp.raise_for_status()
    data = resp.json()
    if data.get("status") == "error":
        return []
    return [int(v["id"]) for k, v in data.items() if k.isdigit() and v.get("active")]


def copiar_portales(id_property_nueva, portales):
    """Copia los portales activos a la nueva propiedad."""
    if not portales:
        return
    url = f"{BASE_URL}/property/update/{id_property_nueva}"
    payload = credenciales()
    for i, pid in enumerate(portales):
        payload[f"portals[{i}]"] = pid
    resp = requests.post(url, data=payload)
    resp.raise_for_status()
    resultado = resp.json()
    if resultado.get("status") == "error":
        raise Exception(f"Error copiando portales: {resultado.get('message')}")


def get_comision(id_property):
    """Lee la comisión empresa y negocio externo de la propiedad original."""
    url = f"{BASE_URL}/property/get-commission/{id_property}"
    resp = requests.get(url, params=credenciales())
    resp.raise_for_status()
    data = resp.json()
    if data.get("status") == "error":
        return None
    return data

def copiar_comision(id_property_nueva, comision):
    """Copia la comisión a la nueva propiedad."""
    url = f"{BASE_URL}/property/update-commission/{id_property_nueva}"
    campos = {
        "commission_type":          comision.get("commission_type", "percent"),
        "commission_value":         comision.get("commission_value", "0"),
        "external_commission_type": comision.get("external_commission_type", "percent"),
        "external_commission_value":comision.get("external_commission_value", "0"),
        "external_commission_active": comision.get("external_commission_active", False),
        "external_commission_note": comision.get("external_commission_note", ""),
        "exclusivity_contract":     comision.get("exclusivity_contract", False),
    }
    if comision.get("exclusivity_contract_date"):
        campos["exclusivity_contract_date"] = comision["exclusivity_contract_date"][:10]
    data = {**credenciales(), **campos}
    resp = requests.post(url, data=data)
    resp.raise_for_status()
    resultado = resp.json()
    if resultado.get("status") == "error":
        raise Exception(f"Error copiando comision: {resultado.get('message')}")
    return resultado

def crear_propiedad(datos, reintentos=3):
    url = f"{BASE_URL}/property/add"
    payload = credenciales()
    for campo, valor in datos.items():
        if campo not in CAMPOS_EXCLUIR and valor not in (None, "", [], {}):
            if isinstance(valor, (dict, list)):
                continue
            payload[campo] = valor
    payload["id_status_on_page"] = "1"
    for intento in range(1, reintentos + 1):
        resp = requests.post(url, data=payload, timeout=30)
        if resp.status_code in (502, 503, 504):
            if intento < reintentos:
                print(f"      ⚠ Servidor no disponible ({resp.status_code}), reintentando ({intento}/{reintentos})...")
                time.sleep(4)
                continue
        resp.raise_for_status()
        resultado = resp.json()
        if resultado.get("status") == "error":
            raise Exception(f"Error al crear: {resultado.get('message')}")
        return resultado

def post_update(id_property, campos):
    url = f"{BASE_URL}/property/update/{id_property}"
    data = {**credenciales(), **campos}
    resp = requests.post(url, data=data)
    resp.raise_for_status()
    resultado = resp.json()
    if resultado.get("status") == "error":
        raise Exception(f"Error update: {resultado.get('message')}")
    return resultado

def copiar_caracteristicas(id_property_nueva, features):
    """Copia características internas y externas a la nueva propiedad."""
    if not features:
        return 0
    ids = []
    for grupo in ("internal", "external"):
        for item in features.get(grupo, []):
            fid = item.get("id")
            if fid:
                ids.append(str(fid))
    if not ids:
        return 0
    url = f"{BASE_URL}/property/update/{id_property_nueva}"
    payload = credenciales()
    for i, fid in enumerate(ids):
        payload[f"features[{i}]"] = fid
    resp = requests.post(url, data=payload)
    resp.raise_for_status()
    resultado = resp.json()
    if resultado.get("status") == "error":
        raise Exception(f"Error copiando características: {resultado.get('message')}")
    return len(ids)

def subir_imagen(id_property, url_imagen, posicion, reintentos=3):
    img_resp = requests.get(url_imagen, timeout=30)
    img_resp.raise_for_status()
    filename = url_imagen.split("/")[-1]
    ext = filename.split(".")[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "gif", "webp"):
        ext = "jpg"
    with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
        tmp.write(img_resp.content)
        tmp_path = tmp.name
    try:
        upload_url = f"{BASE_URL}/property/upload-image/{id_property}"
        params = {"id_company": ID_COMPANY, "wasi_token": WASI_TOKEN}
        for intento in range(1, reintentos + 1):
            with open(tmp_path, "rb") as f:
                files = {"image": (filename, f, f"image/{ext}")}
                resp = requests.post(upload_url, params=params, files=files, timeout=60)
            resp.raise_for_status()
            resultado = resp.json()
            if resultado.get("status") == "success":
                return True
            if intento < reintentos:
                time.sleep(2)
        return False
    finally:
        os.unlink(tmp_path)

def copiar_imagenes(id_property_nueva, galleries):
    if not galleries:
        return 0, 0
    imagenes = []
    for gallery in galleries:
        for key, val in gallery.items():
            if key == "id":
                continue
            if isinstance(val, dict) and "url_original" in val:
                imagenes.append(val)
    imagenes.sort(key=lambda x: x.get("position", 0))
    total = len(imagenes)
    exitosas = 0
    for img in imagenes:
        pos = img.get("position", exitosas + 1)
        url_original = img.get("url_original")
        filename = img.get("filename", "imagen")
        try:
            ok = subir_imagen(id_property_nueva, url_original, pos)
            if ok:
                exitosas += 1
                print(f"        ✓ Imagen {exitosas}/{total}: {filename}")
            else:
                print(f"        ✗ Falló imagen {pos}: {filename}")
        except Exception as e:
            print(f"        ✗ Error imagen {pos}: {e}")
        time.sleep(0.5)
    return exitosas, total

def leer_excel(archivo):
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # Columnas obligatorias
    try:
        col_id  = headers.index("Disponible (codigo)") + 1
        col_com = headers.index("Comentario") + 1
    except ValueError as e:
        raise Exception(f"No se encontró la columna: {e}. Columnas: {headers}")

    propiedades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_prop    = row[col_id - 1]
        comentario = row[col_com - 1]
        if id_prop:
            propiedades.append((
                str(int(id_prop)),
                str(comentario) if comentario else "",
            ))
    return propiedades

def procesar_propiedad(id_property, comentario):
    # Paso 1 - Obtener datos de la propiedad
    propiedad = get_propiedad(id_property)
    titulo    = propiedad.get("title", id_property)
    galleries = propiedad.get("galleries", [])
    features  = propiedad.get("features")
    print(f"      ✓ Encontrada: {titulo}")

    # Obtener portales activos de la original
    print(f"      → Obteniendo portales...")
    portales = get_portales(id_property)
    print(f"      ✓ Portales: {portales}")

    # Paso 2 - Obtener propietario
    print(f"      → Obteniendo propietario...")
    propietario = get_propietario(id_property)
    if propietario:
        id_client = propietario.get("id_client")
        nombre    = propietario.get("first_name", "")
        apellido  = propietario.get("last_name", "")
        telefono  = propietario.get("cell_phone") or propietario.get("phone", "Sin telefono")
        print(f"      ✓ Propietario: {nombre} {apellido} | Tel: {telefono} | ID: {id_client}")
    else:
        id_client = None
        print(f"      ⚠ Sin propietario asignado")

    # Paso 3 - Duplicar propiedad
    print(f"      → Duplicando...")
    nueva    = crear_propiedad(propiedad)
    nuevo_id = nueva.get("id_property", "N/A")
    print(f"      ✓ Nueva propiedad ID: {nuevo_id}")

    # Paso 4 - Activar nueva
    print(f"      → Activando nueva propiedad...")
    post_update(nuevo_id, {"id_status_on_page": "1", "id_availability": "1"})
    print(f"      ✓ Nueva propiedad activa y disponible")
    time.sleep(1)

    # Paso 5 - Asignar propietario
    if id_client:
        print(f"      → Asignando propietario...")
        try:
            asignar_propietario(nuevo_id, id_client)
            print(f"      ✓ Propietario asignado correctamente")
        except Exception as e:
            print(f"      ⚠ No se pudo asignar propietario: {e}")
    time.sleep(1)

    # Paso 6 - Copiar comisiones
    print(f"      → Copiando comisiones...")
    comision = get_comision(id_property)
    if comision:
        try:
            copiar_comision(nuevo_id, comision)
            tipo_emp  = comision.get("commission_type", "")
            val_emp   = comision.get("commission_value", "0")
            tipo_ext  = comision.get("external_commission_type", "")
            val_ext   = comision.get("external_commission_value", "0")
            activa    = comision.get("external_commission_active", False)
            print(f"      ✓ Empresa: {val_emp} ({tipo_emp}) | Externo: {val_ext} ({tipo_ext}) activa={activa}")
        except Exception as e:
            print(f"      ⚠ No se pudieron copiar comisiones: {e}")
    else:
        print(f"      ℹ Sin comisiones registradas")
    time.sleep(1)

    # Paso 7 - Copiar características
    if features:
        print(f"      → Copiando caracteristicas...")
        try:
            n = copiar_caracteristicas(nuevo_id, features)
            print(f"      ✓ {n} caracteristicas copiadas")
        except Exception as e:
            print(f"      ⚠ No se pudieron copiar caracteristicas: {e}")
    else:
        print(f"      ℹ Sin caracteristicas registradas")
    time.sleep(1)

    # Paso 8 - Copiar imágenes
    print(f"      → Copiando imagenes...")
    exitosas, total = copiar_imagenes(nuevo_id, galleries)
    print(f"      ✓ {exitosas}/{total} imagenes copiadas")
    time.sleep(1)

    # Paso 9 - Copiar portales
    if portales:
        print(f"      → Copiando portales ({len(portales)})...")
        try:
            copiar_portales(nuevo_id, portales)
            print(f"      ✓ Portales copiados: {portales}")
        except Exception as e:
            print(f"      ⚠ No se pudieron copiar portales: {e}")
    else:
        print(f"      ℹ Sin portales registrados")
    time.sleep(1)

    # Paso 10 - Agregar observacion a la nueva propiedad
    from datetime import datetime
    fecha = datetime.now().strftime('%Y-%m-%d')
    obs_nueva = f"Disponible - Confirmado con propietario el {fecha}. Cod. nuevo: #{nuevo_id} | Cod. anterior: #{id_property}."
    if comentario:
        obs_nueva += f" {comentario}"
    print(f"      → Agregando observacion a nueva propiedad...")
    try:
        post_update(nuevo_id, {"comment": obs_nueva})
        print(f"      ✓ Observacion: {obs_nueva[:70]}")
    except Exception as e:
        print(f"      ⚠ No se pudo agregar observacion: {e}")
    time.sleep(1)

    # Paso 11 - Inactivar original (NO a papelera — solo vendidas van al basurero)
    print(f"      → Inactivando original (status=2)...")
    nota_original = f"Reemplazada por #{nuevo_id} el {fecha}. {comentario}".strip()
    post_update(id_property, {"comment": nota_original, "id_status_on_page": "2"})
    print(f"      ✓ Original inactivada (no eliminada)")

    return nuevo_id, exitosas, total

def marcar_vendido(id_property: str, comentario: str):
    """Agrega observacion con fecha y envia la propiedad a papelera (status=4)."""
    from datetime import datetime
    nota = f"Vendido el {datetime.now().strftime('%Y-%m-%d')} - Confirmado por propietario. {comentario}".strip()
    post_update(id_property, {"comment": nota, "id_status_on_page": "4"})
    print(f"      ✓ Observacion: {nota[:70]}")
    print(f"      ✓ Enviada a papelera (status=4)")


# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Duplicar propiedades disponibles o marcar como vendidas."
    )
    parser.add_argument("archivo", help="Archivo Excel con columnas 'Disponible (codigo)' y 'Comentario'")
    parser.add_argument("--vendido", action="store_true",
                        help="Modo vendido: agrega observacion con fecha y envia a papelera (NO duplica)")
    args = parser.parse_args()

    modo = "VENDIDO → Papelera" if args.vendido else "Duplicar y Archivar"
    print(f"\n{'='*55}")
    print(f"  WASI - {modo}")
    print(f"{'='*55}")

    try:
        print(f"\nLeyendo archivo: {args.archivo}...")
        propiedades = leer_excel(args.archivo)
        print(f"✓ {len(propiedades)} propiedades encontradas\n")
    except FileNotFoundError:
        print(f"\n✗ No se encontro el archivo: {args.archivo}")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Error leyendo Excel: {e}")
        sys.exit(1)

    exitosos = 0
    fallidos = 0
    errores  = []

    for i, (id_property, comentario) in enumerate(propiedades, 1):
        print(f"\n[{i}/{len(propiedades)}] Propiedad #{id_property}...")
        try:
            if args.vendido:
                marcar_vendido(id_property, comentario)
                print(f"      ─────────────────────────────────────")
                print(f"      #{id_property} → Papelera (vendida)")
            else:
                nuevo_id, imgs_ok, imgs_total = procesar_propiedad(id_property, comentario)
                print(f"      ─────────────────────────────────────")
                print(f"      Original #{id_property} → Inactiva")
                print(f"      Nueva copia → ID #{nuevo_id} (Activa)")
                print(f"      Imagenes   → {imgs_ok}/{imgs_total}")
            exitosos += 1
        except Exception as e:
            print(f"      ✗ Error: {e}")
            errores.append((id_property, str(e)))
            fallidos += 1

        if i < len(propiedades):
            time.sleep(2)

    print(f"\n{'='*55}")
    print(f"  Proceso completado!")
    print(f"  ✓ Exitosos: {exitosos}")
    print(f"  ✗ Fallidos: {fallidos}")
    if errores:
        print(f"\n  Propiedades con error:")
        for id_prop, err in errores:
            print(f"    - #{id_prop}: {err}")
    print(f"{'='*55}\n")

if __name__ == "__main__":
    main()
